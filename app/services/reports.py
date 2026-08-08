from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any
from xml.sax.saxutils import escape

from flask import current_app
from sqlalchemy.orm import joinedload

from app.models import (
    DemandInsight,
    InventoryLocation,
    Product,
    StockLevel,
    utcnow,
)


@dataclass(frozen=True)
class ReportColumn:
    key: str
    header: str
    width: int
    kind: str = "text"
    pdf_width: int = 55


@dataclass
class InventoryReport:
    report_type: str
    title: str
    generated_at: datetime
    methodology: str
    summary: dict[str, Any]
    columns: tuple[ReportColumn, ...]
    rows: list[dict[str, Any]]

    def as_dict(self) -> dict:
        return {
            "report_type": self.report_type,
            "title": self.title,
            "generated_at": self.generated_at.isoformat(),
            "methodology": self.methodology,
            "summary": _json_safe(self.summary),
            "rows": [_json_safe(row) for row in self.rows],
        }


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def _aware(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value


class ReportService:
    """Build deterministic, workspace-scoped report snapshots from live stock."""

    @staticmethod
    def build(report_type: str, *, workspace_id: int) -> InventoryReport:
        if report_type == "risk":
            return ReportService.risk_report(workspace_id=workspace_id)
        if report_type == "valuation":
            return ReportService.valuation_report(workspace_id=workspace_id)
        raise ValueError("report type must be risk or valuation")

    @staticmethod
    def _stock_rows(workspace_id: int, *, active_only: bool) -> list[StockLevel]:
        query = (
            StockLevel.query.options(
                joinedload(StockLevel.product).joinedload(Product.preferred_supplier),
                joinedload(StockLevel.location),
                joinedload(StockLevel.bin),
            )
            .join(Product, StockLevel.product_id == Product.id)
            .join(InventoryLocation, StockLevel.location_id == InventoryLocation.id)
            .filter(InventoryLocation.workspace_id == workspace_id)
        )
        if active_only:
            query = query.filter(Product.is_active.is_(True))
        return query.order_by(Product.name, InventoryLocation.code, StockLevel.id).all()

    @staticmethod
    def risk_report(*, workspace_id: int, now: datetime | None = None) -> InventoryReport:
        now = _aware(now) or utcnow()
        stock_rows = ReportService._stock_rows(workspace_id, active_only=True)
        grouped: dict[tuple[int, int], list[StockLevel]] = defaultdict(list)
        for stock in stock_rows:
            grouped[(stock.product_id, stock.location_id)].append(stock)

        location_ids = {location_id for _, location_id in grouped}
        insights_query = DemandInsight.query
        if location_ids:
            insights_query = insights_query.filter(
                DemandInsight.location_id.in_(location_ids)
            )
        else:
            insights_query = insights_query.filter(False)
        latest: dict[tuple[int, int], DemandInsight] = {}
        for insight in insights_query.order_by(
            DemandInsight.generated_at.desc(), DemandInsight.id.desc()
        ):
            latest.setdefault((insight.product_id, insight.location_id), insight)

        critical_days = current_app.config["CRITICAL_STOCKOUT_DAYS"]
        rows: list[dict[str, Any]] = []
        for key, positions in grouped.items():
            product = positions[0].product
            location = positions[0].location
            insight = latest.get(key)
            on_hand = sum(
                (Decimal(row.quantity_on_hand or 0) for row in positions),
                start=Decimal("0.00"),
            )
            reserved = sum(
                (Decimal(row.quantity_reserved or 0) for row in positions),
                start=Decimal("0.00"),
            )
            available = on_hand - reserved
            daily_demand = Decimal(str(insight.daily_demand if insight else 0)).quantize(
                Decimal("0.01")
            )
            reorder = Decimal(
                insight.recommended_reorder_quantity if insight else 0
            ).quantize(Decimal("0.01"))
            stockout_at = _aware(insight.expected_stockout_at) if insight else None
            days_cover = (
                (available / daily_demand).quantize(Decimal("0.01"))
                if daily_demand > 0
                else None
            )
            if available <= 0 or (
                stockout_at and stockout_at <= now + timedelta(days=critical_days)
            ):
                risk = "critical"
            elif reorder > 0 or available <= Decimal(product.reorder_point or 0):
                risk = "watch"
            else:
                risk = "healthy"

            if insight and insight.narrative:
                action = insight.narrative
            elif risk == "critical":
                action = "Replenish immediately and review open reservations."
            elif risk == "watch":
                action = "Review demand and create a replenishment order."
            else:
                action = "No replenishment action is currently required."

            supplier = product.preferred_supplier
            rows.append(
                {
                    "risk": risk,
                    "sku": product.sku,
                    "product": product.name,
                    "category": product.category,
                    "location": location.code,
                    "quantity_on_hand": on_hand,
                    "quantity_reserved": reserved,
                    "quantity_available": available,
                    "daily_demand": daily_demand,
                    "days_of_cover": days_cover,
                    "projected_stockout": stockout_at,
                    "reorder_quantity": reorder,
                    "supplier": supplier.name if supplier else None,
                    "lead_time_days": (
                        supplier.lead_time_days
                        if supplier
                        else current_app.config["DEFAULT_SUPPLIER_LEAD_TIME_DAYS"]
                    ),
                    "action": action,
                }
            )

        risk_order = {"critical": 0, "watch": 1, "healthy": 2}
        rows.sort(key=lambda row: (risk_order[row["risk"]], row["product"], row["location"]))
        summary = {
            "stock_positions": len(rows),
            "critical": sum(row["risk"] == "critical" for row in rows),
            "watch": sum(row["risk"] == "watch" for row in rows),
            "healthy": sum(row["risk"] == "healthy" for row in rows),
            "recommended_reorder_units": sum(
                (Decimal(row["reorder_quantity"]) for row in rows),
                start=Decimal("0.00"),
            ),
        }
        return InventoryReport(
            report_type="risk",
            title="Inventory Risk and Reorder Report",
            generated_at=now,
            methodology=(
                "Uses authoritative available stock (on hand minus reserved) and the latest "
                "forecast for each product/location. Critical means out of stock or projected "
                f"stockout within {critical_days} day(s)."
            ),
            summary=summary,
            columns=(
                ReportColumn("risk", "Risk", 12, pdf_width=42),
                ReportColumn("sku", "SKU", 18, pdf_width=58),
                ReportColumn("product", "Product", 28, pdf_width=90),
                ReportColumn("category", "Category", 18, pdf_width=65),
                ReportColumn("location", "Location", 14, pdf_width=50),
                ReportColumn("quantity_on_hand", "On hand", 13, "number", 44),
                ReportColumn("quantity_reserved", "Reserved", 13, "number", 44),
                ReportColumn("quantity_available", "Available", 13, "number", 44),
                ReportColumn("daily_demand", "Daily demand", 15, "number", 48),
                ReportColumn("days_of_cover", "Days cover", 13, "number", 45),
                ReportColumn("reorder_quantity", "Reorder", 13, "number", 45),
                ReportColumn("supplier", "Supplier", 24, pdf_width=80),
                ReportColumn("action", "Recommended action", 46, pdf_width=120),
            ),
            rows=rows,
        )

    @staticmethod
    def valuation_report(*, workspace_id: int, now: datetime | None = None) -> InventoryReport:
        now = _aware(now) or utcnow()
        stock_rows = ReportService._stock_rows(workspace_id, active_only=False)
        rows: list[dict[str, Any]] = []
        for stock in stock_rows:
            on_hand = Decimal(stock.quantity_on_hand or 0)
            reserved = Decimal(stock.quantity_reserved or 0)
            if on_hand == 0 and reserved == 0:
                continue
            available = on_hand - reserved
            unit_cost = Decimal(stock.product.cost_price or 0)
            supplier = stock.product.preferred_supplier
            rows.append(
                {
                    "sku": stock.product.sku,
                    "product": stock.product.name,
                    "category": stock.product.category,
                    "location": stock.location.code,
                    "bin": stock.bin.code if stock.bin else "Unassigned",
                    "quantity_on_hand": on_hand,
                    "quantity_reserved": reserved,
                    "quantity_available": available,
                    "unit_cost": unit_cost,
                    "on_hand_value": (on_hand * unit_cost).quantize(Decimal("0.01")),
                    "available_value": (available * unit_cost).quantize(Decimal("0.01")),
                    "supplier": supplier.name if supplier else None,
                    "product_status": "active" if stock.product.is_active else "archived",
                }
            )

        rows.sort(key=lambda row: (row["product"], row["location"], row["bin"]))
        summary = {
            "sku_count": len({row["sku"] for row in rows}),
            "stock_positions": len(rows),
            "quantity_on_hand": sum(
                (Decimal(row["quantity_on_hand"]) for row in rows),
                start=Decimal("0.00"),
            ),
            "quantity_reserved": sum(
                (Decimal(row["quantity_reserved"]) for row in rows),
                start=Decimal("0.00"),
            ),
            "quantity_available": sum(
                (Decimal(row["quantity_available"]) for row in rows),
                start=Decimal("0.00"),
            ),
            "on_hand_value": sum(
                (Decimal(row["on_hand_value"]) for row in rows),
                start=Decimal("0.00"),
            ),
            "available_value": sum(
                (Decimal(row["available_value"]) for row in rows),
                start=Decimal("0.00"),
            ),
            "currency": current_app.config["REPORT_CURRENCY"],
        }
        return InventoryReport(
            report_type="valuation",
            title="Inventory Valuation Report",
            generated_at=now,
            methodology=(
                "Current-unit-cost valuation: product cost_price multiplied by the "
                "authoritative quantity on hand. This is not FIFO, LIFO, or weighted-average "
                "valuation because receipt-level cost layers are not yet recorded."
            ),
            summary=summary,
            columns=(
                ReportColumn("sku", "SKU", 18, pdf_width=58),
                ReportColumn("product", "Product", 28, pdf_width=105),
                ReportColumn("category", "Category", 18, pdf_width=62),
                ReportColumn("location", "Location", 14, pdf_width=48),
                ReportColumn("bin", "Bin", 14, pdf_width=48),
                ReportColumn("quantity_on_hand", "On hand", 13, "number", 45),
                ReportColumn("quantity_reserved", "Reserved", 13, "number", 45),
                ReportColumn("quantity_available", "Available", 13, "number", 45),
                ReportColumn("unit_cost", "Unit cost", 14, "currency", 55),
                ReportColumn("on_hand_value", "On-hand value", 17, "currency", 67),
                ReportColumn("available_value", "Available value", 17, "currency", 67),
                ReportColumn("supplier", "Supplier", 24, pdf_width=90),
                ReportColumn("product_status", "Status", 12, pdf_width=46),
            ),
            rows=rows,
        )


class ReportExporter:
    """Render a report snapshot into portable Excel or PDF bytes."""

    @staticmethod
    def export(report: InventoryReport, file_format: str) -> tuple[bytes, str]:
        normalized = file_format.strip().lower()
        if normalized == "xlsx":
            return ReportExporter.to_xlsx(report), (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        if normalized == "pdf":
            return ReportExporter.to_pdf(report), "application/pdf"
        raise ValueError("report format must be xlsx or pdf")

    @staticmethod
    def filename(report: InventoryReport, file_format: str) -> str:
        date = report.generated_at.strftime("%Y-%m-%d")
        return f"stockpilot-{report.report_type}-{date}.{file_format.lower()}"

    @staticmethod
    def to_xlsx(report: InventoryReport) -> bytes:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        detail_sheet = workbook.create_sheet("Details")
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        navy = "14213D"
        blue = "1D4ED8"
        pale_blue = "E8F0FE"
        pale_gray = "F4F6F8"
        white = "FFFFFF"
        thin_gray = Side(style="thin", color="D9DEE7")

        summary_sheet.merge_cells("A1:D1")
        title = summary_sheet["A1"]
        title.value = report.title
        title.font = Font(name="Aptos Display", size=18, bold=True, color=white)
        title.fill = PatternFill("solid", fgColor=navy)
        title.alignment = Alignment(vertical="center")
        summary_sheet.row_dimensions[1].height = 34
        summary_sheet["A3"] = "Generated"
        summary_sheet["B3"] = report.generated_at.replace(tzinfo=None)
        summary_sheet["B3"].number_format = "yyyy-mm-dd hh:mm"
        summary_sheet["A4"] = "Methodology"
        summary_sheet.merge_cells("A4:A5")
        summary_sheet["A4"].alignment = Alignment(vertical="center")
        summary_sheet.merge_cells("B4:D5")
        summary_sheet["B4"] = report.methodology
        summary_sheet["B4"].alignment = Alignment(wrap_text=True, vertical="center")
        summary_sheet.row_dimensions[4].height = 28
        summary_sheet.row_dimensions[5].height = 28

        row_index = 7
        summary_sheet.cell(row=row_index, column=1, value="Metric")
        summary_sheet.cell(row=row_index, column=2, value="Value")
        for cell in summary_sheet[row_index][:2]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=blue)
        for label, value in report.summary.items():
            row_index += 1
            summary_sheet.cell(
                row=row_index, column=1, value=_summary_label(label)
            )
            cell = summary_sheet.cell(row=row_index, column=2, value=_excel_value(value))
            if "value" in label:
                cell.number_format = '#,##0.00'
            elif isinstance(value, (int, float, Decimal)):
                cell.number_format = (
                    '#,##0'
                    if label
                    in {
                        "sku_count",
                        "stock_positions",
                        "critical",
                        "watch",
                        "healthy",
                    }
                    else '#,##0.00'
                )
        summary_sheet.column_dimensions["A"].width = 30
        summary_sheet.column_dimensions["B"].width = 30
        summary_sheet.column_dimensions["C"].width = 25
        summary_sheet.column_dimensions["D"].width = 25
        summary_sheet.sheet_view.showGridLines = False
        summary_sheet.freeze_panes = "A7"

        detail_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report.columns))
        detail_title = detail_sheet.cell(row=1, column=1, value=report.title)
        detail_title.font = Font(name="Aptos Display", size=16, bold=True, color=white)
        detail_title.fill = PatternFill("solid", fgColor=navy)
        detail_title.alignment = Alignment(vertical="center")
        detail_sheet.row_dimensions[1].height = 30
        detail_sheet.cell(
            row=2,
            column=1,
            value=f"Snapshot generated {report.generated_at.strftime('%Y-%m-%d %H:%M %Z')}",
        )
        header_row = 4
        for column_index, column in enumerate(report.columns, start=1):
            cell = detail_sheet.cell(row=header_row, column=column_index, value=column.header)
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = Border(bottom=thin_gray)
            detail_sheet.column_dimensions[cell.column_letter].width = column.width

        for row_index, row in enumerate(report.rows, start=header_row + 1):
            for column_index, column in enumerate(report.columns, start=1):
                value = row.get(column.key)
                cell = detail_sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=_excel_value(value),
                )
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=column.key in {"action", "product"},
                )
                if column.kind == "currency":
                    cell.number_format = '#,##0.00'
                elif column.kind == "number":
                    cell.number_format = '#,##0.00'
                elif column.kind == "datetime":
                    cell.number_format = "yyyy-mm-dd hh:mm"
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=pale_gray)

        if report.rows:
            end_row = header_row + len(report.rows)
            end_column_letter = detail_sheet.cell(
                row=header_row, column=len(report.columns)
            ).column_letter
            table = Table(
                displayName=f"{report.report_type.title()}ReportTable",
                ref=f"A{header_row}:{end_column_letter}{end_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            detail_sheet.add_table(table)

            if report.report_type == "risk":
                risk_column = next(
                    index
                    for index, column in enumerate(report.columns, start=1)
                    if column.key == "risk"
                )
                risk_letter = detail_sheet.cell(
                    row=header_row + 1, column=risk_column
                ).column_letter
                data_range = f"A{header_row + 1}:{end_column_letter}{end_row}"
                detail_sheet.conditional_formatting.add(
                    data_range,
                    FormulaRule(
                        formula=[f'${risk_letter}{header_row + 1}="critical"'],
                        fill=PatternFill("solid", fgColor="FDE8E8"),
                    ),
                )
                detail_sheet.conditional_formatting.add(
                    data_range,
                    FormulaRule(
                        formula=[f'${risk_letter}{header_row + 1}="watch"'],
                        fill=PatternFill("solid", fgColor="FFF4D6"),
                    ),
                )

        detail_sheet.freeze_panes = f"A{header_row + 1}"
        detail_sheet.auto_filter.ref = detail_sheet.dimensions
        detail_sheet.sheet_view.showGridLines = False
        detail_sheet.sheet_properties.pageSetUpPr.fitToPage = True
        detail_sheet.page_setup.fitToWidth = 1
        detail_sheet.page_setup.fitToHeight = 0
        detail_sheet.sheet_view.zoomScale = 90
        for cell in summary_sheet[3] + summary_sheet[4]:
            if cell.value is not None:
                cell.fill = PatternFill("solid", fgColor=pale_blue)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    @staticmethod
    def to_pdf(report: InventoryReport) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        stream = BytesIO()
        document = SimpleDocTemplate(
            stream,
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=13 * mm,
            title=report.title,
            author="StockPilot AI",
        )
        styles = getSampleStyleSheet()
        styles.add(
            ParagraphStyle(
                name="ReportTitle",
                parent=styles["Title"],
                fontName="Helvetica-Bold",
                fontSize=18,
                leading=22,
                textColor=colors.HexColor("#14213D"),
                alignment=TA_LEFT,
                spaceAfter=6,
            )
        )
        styles.add(
            ParagraphStyle(
                name="ReportNote",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=8,
                leading=11,
                textColor=colors.HexColor("#44546A"),
            )
        )
        styles.add(
            ParagraphStyle(
                name="TableCell",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=6.4,
                leading=8,
            )
        )

        story: list[Any] = [
            Paragraph(escape(report.title), styles["ReportTitle"]),
            Paragraph(
                f"Generated {report.generated_at.strftime('%Y-%m-%d %H:%M %Z')}",
                styles["ReportNote"],
            ),
            Spacer(1, 5),
        ]
        summary_data = [["Metric", "Value"]] + [
            [_summary_label(label), _display_value(value, report)]
            for label, value in report.summary.items()
        ]
        summary_table = Table(summary_data, colWidths=[55 * mm, 48 * mm], hAlign="LEFT")
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6F8")]),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9DEE7")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.extend(
            [
                summary_table,
                Spacer(1, 7),
                Paragraph("Methodology: " + escape(report.methodology), styles["ReportNote"]),
                Spacer(1, 9),
            ]
        )

        headers = [Paragraph(escape(column.header), styles["TableCell"]) for column in report.columns]
        table_data: list[list[Any]] = [headers]
        for row in report.rows:
            table_data.append(
                [
                    Paragraph(
                        escape(_display_value(row.get(column.key), report)),
                        styles["TableCell"],
                    )
                    for column in report.columns
                ]
            )
        if not report.rows:
            table_data.append(
                [
                    Paragraph("No report rows matched this snapshot.", styles["TableCell"])
                ]
                + [""] * (len(report.columns) - 1)
            )

        detail_table = Table(
            table_data,
            colWidths=[column.pdf_width for column in report.columns],
            repeatRows=1,
            hAlign="LEFT",
        )
        detail_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14213D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9DEE7")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        if report.report_type == "risk":
            risk_column = next(
                index for index, column in enumerate(report.columns) if column.key == "risk"
            )
            for row_number, row in enumerate(report.rows, start=1):
                color = (
                    colors.HexColor("#FDE8E8")
                    if row["risk"] == "critical"
                    else colors.HexColor("#FFF4D6")
                    if row["risk"] == "watch"
                    else None
                )
                if color:
                    detail_styles.append(("BACKGROUND", (0, row_number), (-1, row_number), color))
                detail_styles.append(
                    ("FONTNAME", (risk_column, row_number), (risk_column, row_number), "Helvetica-Bold")
                )
        detail_table.setStyle(TableStyle(detail_styles))
        story.append(detail_table)

        def footer(canvas, doc):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#667085"))
            canvas.drawString(10 * mm, 7 * mm, "StockPilot AI - operational report snapshot")
            canvas.drawRightString(
                landscape(A4)[0] - 10 * mm,
                7 * mm,
                f"Page {doc.page}",
            )
            canvas.restoreState()

        document.build(story, onFirstPage=footer, onLaterPages=footer)
        return stream.getvalue()


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def _display_value(value: Any, report: InventoryReport) -> str:
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, Decimal):
        return f"{value:,.2f}"
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)


def _summary_label(key: str) -> str:
    labels = {
        "sku_count": "SKU count",
        "stock_positions": "Stock positions",
        "quantity_on_hand": "Quantity on hand",
        "quantity_reserved": "Quantity reserved",
        "quantity_available": "Quantity available",
        "on_hand_value": "On-hand value",
        "available_value": "Available value",
        "recommended_reorder_units": "Recommended reorder units",
    }
    return labels.get(key, key.replace("_", " ").title())
