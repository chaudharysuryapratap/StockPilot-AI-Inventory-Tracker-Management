from __future__ import annotations

import sqlite3
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import inspect

from app import create_app, db
from app.models import (
    AlertDelivery,
    DemandInsight,
    Product,
    StockLevel,
    Supplier,
    User,
    Workspace,
    utcnow,
)
from app.schema import SPRINT_6_SCHEMA_VERSION, current_schema_versions, migrate_schema
from app.services.emailer import ReportMailer
from app.services.forecast import ForecastResult
from app.services.reports import ReportService


INTERNAL_HEADERS = {"X-Internal-Token": "test-internal-token"}


def test_supplier_api_crud_validation_archive_and_product_link(
    client, app, seeded_catalog
):
    created = client.post(
        "/api/suppliers",
        headers=INTERNAL_HEADERS,
        json={
            "name": "Noida Fresh Foods",
            "contact_email": "orders@noidafresh.test",
            "contact_phone": "+91 99999 11111",
            "lead_time_days": 2,
            "payment_terms": "Net 15",
        },
    )
    assert created.status_code == 201
    supplier_id = created.json["supplier"]["id"]
    assert created.json["supplier"]["contact_email"] == "orders@noidafresh.test"
    assert created.json["supplier"]["payment_terms"] == "Net 15"

    duplicate = client.post(
        "/api/suppliers",
        headers=INTERNAL_HEADERS,
        json={"name": "noida fresh foods", "lead_time_days": 3},
    )
    invalid = client.post(
        "/api/suppliers",
        headers=INTERNAL_HEADERS,
        json={"name": "Bad supplier", "contact_email": "not-an-email"},
    )
    assert duplicate.status_code == 409
    assert duplicate.json["fields"]["name"] == "already exists in this workspace"
    assert invalid.status_code == 400
    assert "valid email" in invalid.json["fields"]["contact_email"]

    updated = client.patch(
        f"/api/suppliers/{supplier_id}",
        headers=INTERNAL_HEADERS,
        json={"lead_time_days": 5, "payment_terms": "Net 30"},
    )
    assert updated.status_code == 200
    assert updated.json["supplier"]["lead_time_days"] == 5

    archived = client.delete(
        f"/api/suppliers/{supplier_id}", headers=INTERNAL_HEADERS
    )
    assert archived.status_code == 200
    assert archived.json["supplier"]["is_active"] is False
    hidden = client.get("/api/suppliers", headers=INTERNAL_HEADERS)
    assert all(row["id"] != supplier_id for row in hidden.json["suppliers"])

    rejected_product = client.post(
        "/api/products",
        headers=INTERNAL_HEADERS,
        json={
            "sku": "SUPPLIER-LINK-1",
            "name": "Supplier-linked product",
            "preferred_supplier_id": supplier_id,
        },
    )
    assert rejected_product.status_code == 400
    assert rejected_product.json["fields"]["preferred_supplier_id"] == "was not found"

    restored = client.post(
        f"/api/suppliers/{supplier_id}/restore", headers=INTERNAL_HEADERS
    )
    linked_product = client.post(
        "/api/products",
        headers=INTERNAL_HEADERS,
        json={
            "sku": "SUPPLIER-LINK-1",
            "name": "Supplier-linked product",
            "preferred_supplier_id": supplier_id,
        },
    )
    assert restored.status_code == 200
    assert linked_product.status_code == 201
    assert linked_product.json["product"]["supplier"] == "Noida Fresh Foods"


def test_supplier_queries_and_mutations_are_workspace_scoped(client, app):
    with app.app_context():
        other_workspace = Workspace(name="Other workspace")
        db.session.add(other_workspace)
        db.session.flush()
        other_user = User(
            workspace=other_workspace,
            name="Other Admin",
            email="other-admin@test.local",
            role="admin",
            is_active=True,
        )
        other_supplier = Supplier(
            workspace=other_workspace,
            name="Other Supplier",
            lead_time_days=4,
        )
        db.session.add_all([other_user, other_supplier])
        db.session.commit()
        supplier_id = other_supplier.id

    own_listing = client.get(
        "/api/suppliers",
        headers={**INTERNAL_HEADERS, "X-Actor-Email": "other-admin@test.local"},
    )
    cross_workspace_update = client.patch(
        f"/api/suppliers/{supplier_id}",
        headers=INTERNAL_HEADERS,
        json={"lead_time_days": 8},
    )
    assert [row["name"] for row in own_listing.json["suppliers"]] == [
        "Other Supplier"
    ]
    assert cross_workspace_update.status_code == 404


def test_risk_and_current_cost_valuation_reports_use_authoritative_stock(
    client, app, seeded_catalog
):
    with app.app_context():
        product = db.session.get(Product, seeded_catalog["product_id"])
        stock = StockLevel.query.one()
        product.cost_price = Decimal("12.50")
        stock.quantity_on_hand = Decimal("6.00")
        stock.quantity_reserved = Decimal("2.00")
        db.session.add(
            DemandInsight(
                product_id=product.id,
                location_id=stock.location_id,
                daily_demand=Decimal("3.00"),
                expected_stockout_at=utcnow() + timedelta(days=1),
                recommended_reorder_quantity=Decimal("8.00"),
                confidence=80,
                narrative="Order 8 units before tomorrow's rush.",
            )
        )
        db.session.commit()
        workspace_id = User.query.first().workspace_id

        risk = ReportService.risk_report(workspace_id=workspace_id)
        valuation = ReportService.valuation_report(workspace_id=workspace_id)

    assert risk.summary["critical"] == 1
    assert risk.rows[0]["quantity_available"] == Decimal("4.00")
    assert risk.rows[0]["reorder_quantity"] == Decimal("8.00")
    assert valuation.summary["on_hand_value"] == Decimal("75.00")
    assert valuation.summary["available_value"] == Decimal("50.00")
    assert "not FIFO, LIFO, or weighted-average" in valuation.methodology

    api_response = client.get(
        "/api/reports/valuation", headers=INTERNAL_HEADERS
    )
    assert api_response.status_code == 200
    assert api_response.json["report"]["summary"]["on_hand_value"] == 75


def test_excel_report_download_contains_summary_and_typed_details(
    client, app, seeded_catalog
):
    with app.app_context():
        product = db.session.get(Product, seeded_catalog["product_id"])
        product.cost_price = Decimal("20.00")
        db.session.commit()

    response = client.get(
        "/api/reports/valuation?format=xlsx", headers=INTERNAL_HEADERS
    )
    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "stockpilot-valuation-" in response.headers["Content-Disposition"]

    workbook = load_workbook(BytesIO(response.data), data_only=False)
    assert workbook.sheetnames == ["Summary", "Details"]
    assert workbook["Summary"]["A1"].value == "Inventory Valuation Report"
    headers = [cell.value for cell in workbook["Details"][4]]
    assert "On-hand value" in headers
    value_column = headers.index("On-hand value") + 1
    assert workbook["Details"].cell(row=5, column=value_column).value == 200
    assert workbook["Details"].freeze_panes == "A5"


def test_pdf_report_download_is_a_nonempty_multipage_safe_document(
    client, seeded_catalog
):
    response = client.get("/api/reports/risk?format=pdf", headers=INTERNAL_HEADERS)
    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert response.data.startswith(b"%PDF-")
    assert len(response.data) > 2500
    assert b"/Type /Page" in response.data


def test_critical_email_delivery_is_sent_and_audited(app, seeded_catalog):
    result = ForecastResult(
        product_id=seeded_catalog["product_id"],
        product_sku="TEST-001",
        product_name="Test Product",
        location_id=seeded_catalog["location_id"],
        location_code="TEST",
        current_stock=1,
        daily_demand=4,
        expected_stockout_at=utcnow() + timedelta(hours=6),
        recommended_reorder_quantity=12,
        confidence=88,
        narrative="Order 12 units immediately.",
    )
    with app.app_context():
        actor = User.query.first()
        app.config.update(
            SES_ENABLED=True,
            SES_FROM_EMAIL="alerts@stockpilot.test",
            ALERT_RECIPIENTS=["owner@freshmart.test", "owner@freshmart.test"],
        )
        mocked_client = type(
            "SESClient",
            (),
            {"send_email": lambda self, **kwargs: {"MessageId": "ses-message-1"}},
        )()
        with patch("app.services.emailer.boto3.client", return_value=mocked_client) as factory:
            delivery = ReportMailer.send_critical_alerts(
                [result], workspace_id=actor.workspace_id
            )

        saved = AlertDelivery.query.one()
        assert delivery.sent is True
        assert delivery.provider_message_id == "ses-message-1"
        assert saved.status == "sent"
        assert saved.recipient_count == 1
        assert saved.item_count == 1
        assert saved.sent_at is not None
        factory.assert_called_once_with("sesv2", region_name="ap-south-1")


def test_noncritical_or_disabled_email_is_skipped_but_recorded(app, seeded_catalog):
    healthy = ForecastResult(
        product_id=seeded_catalog["product_id"],
        product_sku="TEST-001",
        product_name="Test Product",
        location_id=seeded_catalog["location_id"],
        location_code="TEST",
        current_stock=100,
        daily_demand=1,
        expected_stockout_at=utcnow() + timedelta(days=100),
        recommended_reorder_quantity=0,
        confidence=90,
        narrative="Adequately covered.",
    )
    with app.app_context():
        actor = User.query.first()
        with patch("app.services.emailer.boto3.client") as factory:
            result = ReportMailer.send_critical_alerts(
                [healthy], workspace_id=actor.workspace_id
            )
        assert result.sent is False
        assert "No critical" in result.reason
        assert AlertDelivery.query.one().status == "skipped"
        factory.assert_not_called()


def test_sprint5_migrates_legacy_supplier_columns_without_losing_contact(tmp_path):
    database_path = tmp_path / "sprint4-supplier.db"
    connection = sqlite3.connect(database_path)
    connection.executescript(
        """
        CREATE TABLE suppliers (
            id INTEGER PRIMARY KEY,
            name VARCHAR(120) NOT NULL UNIQUE,
            email VARCHAR(255),
            phone VARCHAR(40),
            lead_time_days INTEGER NOT NULL,
            created_at DATETIME NOT NULL
        );
        INSERT INTO suppliers
            (id, name, email, phone, lead_time_days, created_at)
        VALUES
            (1, 'Legacy Foods', 'legacy@example.com', '0123456789', 6,
             '2026-01-01 00:00:00');
        """
    )
    connection.commit()
    connection.close()

    migration_app = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": f"sqlite:///{database_path}",
            "BEDROCK_ENABLED": False,
            "SES_ENABLED": False,
            "STAFF_AUTH_ENABLED": False,
        }
    )
    with migration_app.app_context():
        result = migrate_schema()
        supplier = db.session.get(Supplier, 1)
        columns = {
            column["name"] for column in inspect(db.engine).get_columns("suppliers")
        }
        tables = set(inspect(db.engine).get_table_names())

        assert result.version == SPRINT_6_SCHEMA_VERSION
        assert supplier.name == "Legacy Foods"
        assert supplier.contact_email == "legacy@example.com"
        assert supplier.contact_phone == "0123456789"
        assert supplier.workspace_id is not None
        assert supplier.is_active is True
        assert {"workspace_id", "contact_email", "contact_phone", "payment_terms"}.issubset(
            columns
        )
        assert "alert_deliveries" in tables
        assert SPRINT_6_SCHEMA_VERSION in current_schema_versions()
        db.session.remove()


def test_supplier_and_report_pages_load_and_picker_is_blocked(
    client, app, seeded_catalog
):
    suppliers = client.get("/suppliers")
    reports = client.get("/reports")
    assert suppliers.status_code == 200
    assert b"Supplier records" in suppliers.data
    assert reports.status_code == 200
    assert b"Inventory valuation report" in reports.data

    with app.app_context():
        actor = User.query.first()
        picker = User(
            workspace_id=actor.workspace_id,
            name="Report Picker",
            email="report-picker@test.local",
            role="picker",
            is_active=True,
        )
        db.session.add(picker)
        db.session.commit()
        picker_id = picker.id
    with client.session_transaction() as session:
        session["user_id"] = picker_id
    assert client.get("/reports").status_code == 403
    assert client.get("/suppliers").status_code == 403
